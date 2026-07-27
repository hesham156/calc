"""Excel / CSV export of an employee monthly report."""
import csv
import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.models import Attendance, AttendanceSummary, Employee, WorkSettings
from app.services.calculator import effective_work_window

STATUS_AR = {
    "present": "حاضر", "absent": "غائب", "leave": "إجازة",
    "weekend": "عطلة", "holiday": "عطلة رسمية", "incomplete": "بصمة ناقصة",
}

HEADERS = [
    "التاريخ", "اليوم", "الوردية", "بداية الدوام", "نهاية الدوام", "الدخول", "الخروج", "مدة العمل",
    "التأخير", "انصراف مبكر", "أوفر تايم", "الخصم (دقائق)", "الخصم (مبلغ)", "الحالة",
]


def fmt_minutes(m: int) -> str:
    return f"{m // 60:02d}:{m % 60:02d}"


def fmt_time(t) -> str:
    return t.strftime("%H:%M") if t else "-"


def row_values(a: Attendance, work_settings: WorkSettings) -> list:
    work_start, work_end = effective_work_window(a, work_settings)
    return [
        a.date.strftime("%Y-%m-%d"), a.weekday, a.shift or "-",
        fmt_time(work_start), fmt_time(work_end),
        fmt_time(a.check_in), fmt_time(a.check_out) + ("+1" if a.out_next_day else ""),
        fmt_minutes(a.worked_minutes), fmt_minutes(a.late_minutes),
        fmt_minutes(a.early_leave_minutes), fmt_minutes(a.overtime_minutes),
        a.deduction_minutes, a.deduction_amount, STATUS_AR.get(a.status, a.status),
    ]


def totals_row(rows: list[Attendance]) -> list:
    """Grand-total line closing the detail table."""
    return [
        "الإجمالي", f"{len(rows)} يوم", "", "", "", "", "",
        fmt_minutes(sum(a.worked_minutes for a in rows)),
        fmt_minutes(sum(a.late_minutes for a in rows)),
        fmt_minutes(sum(a.early_leave_minutes for a in rows)),
        fmt_minutes(sum(a.overtime_minutes for a in rows)),
        sum(a.deduction_minutes for a in rows),
        round(sum(a.deduction_amount for a in rows), 2),
        "",
    ]


def summary_rows(s: AttendanceSummary | None) -> list[tuple[str, str]]:
    if not s:
        return []
    return [
        ("إجمالي أيام العمل", str(s.work_days)),
        ("أيام الحضور", str(s.present_days)),
        ("أيام الغياب", str(s.absent_days)),
        ("أيام الإجازة", str(s.leave_days)),
        ("أيام العطل", str(s.weekend_days)),
        ("إجمالي ساعات العمل", fmt_minutes(s.worked_minutes)),
        ("إجمالي التأخير", fmt_minutes(s.late_minutes) + f" ({s.late_minutes} دقيقة)"),
        ("إجمالي الانصراف المبكر", fmt_minutes(s.early_leave_minutes)),
        ("إجمالي الأوفر تايم", fmt_minutes(s.overtime_minutes)),
        ("إجمالي الراحة", fmt_minutes(s.break_minutes)),
        ("إجمالي الخصومات", f"{fmt_minutes(s.deduction_minutes)} = {s.deduction_amount}"),
        ("إجمالي المستحقات (أوفر تايم)", str(s.overtime_amount)),
        ("صافي ساعات العمل", fmt_minutes(s.net_minutes)),
    ]


def export_csv(
    employee: Employee, rows: list[Attendance], summary: AttendanceSummary | None,
    work_settings: WorkSettings,
) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(HEADERS)
    for a in rows:
        w.writerow(row_values(a, work_settings))
    w.writerow(totals_row(rows))
    w.writerow([])
    for label, value in summary_rows(summary):
        w.writerow([label, value])
    return ("﻿" + buf.getvalue()).encode("utf-8")  # BOM so Excel opens Arabic correctly


def export_excel(
    employee: Employee, rows: list[Attendance], summary: AttendanceSummary | None,
    year: int, month: int, work_settings: WorkSettings,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}-{month:02d}"
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    company = work_settings.company_name or "تقرير الحضور والانصراف"
    ws.merge_cells("A1:N1")
    ws["A1"] = company
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center
    ws.merge_cells("A2:N2")
    ws["A2"] = f"الموظف: {employee.name} ({employee.code}) — الشهر: {year}-{month:02d}"
    ws["A2"].alignment = center

    ws.append([])
    ws.append(HEADERS)
    for cell in ws[ws.max_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for a in rows:
        ws.append(row_values(a, work_settings))
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = center

    ws.append(totals_row(rows))
    for cell in ws[ws.max_row]:
        cell.border = border
        cell.alignment = center
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E2E8F0")

    ws.append([])
    for label, value in summary_rows(summary):
        ws.append([label, value])
        ws[ws.max_row][0].font = Font(bold=True)

    ws.append([])
    ws.append(["التوقيع: ______________________", "", "", "", "اعتماد المدير: ______________________"])

    for col, width in zip("ABCDEFGHIJKLMN", [12, 10, 12, 12, 12, 8, 8, 10, 9, 12, 10, 12, 12, 12]):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
